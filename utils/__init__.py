import requests
import openpyxl
from openpyxl import load_workbook


"""
    excel读取写入，样式设置以及爬取页面
"""


def write_excel_xlsx(path, sheet_name, value):
    """
    将数据写入excel,注意是xlsx格式
    :param path: 文件路径
    :param sheet_name: sheet名
    :param value:
    :return:
    """
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))



    workbook.save(path)
    # 设置单元格样式
    set_excel_style(path)
    print("爬取数据已经写入文件！")


# 设置excel文件格式
def set_excel_style(file_path):
    # 调整单元格格式
    wb = load_workbook(file_path)
    ws = wb[wb.sheetnames[0]]

    width = 100
    height = 20

    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = height
    for i in range(1, ws.max_column + 1):
        column = chr(ord('A') - 1 + i)
        if column == 'C':
            ws.column_dimensions[column].width = 16
        elif column == 'B':
            ws.column_dimensions[column].width = 70
        else:
            ws.column_dimensions[column].width = width
    wb.save(file_path)


# 读取excel文件
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()


# 根据链接获取请求页面的内容,这种是通过get的方式获取数据
def getHTMLText(url):
    try:
        r = requests.get(url, timeout=30)
        # 判断是否返回异常
        r.raise_for_status()
        # 从网页的内容中分析网页编码的方式
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ""


# 格式化字符串
def str_format(str):
    if str != None:
        return str.replace('\n', '').replace('\r', '').replace('\t', '');
    else:
        return ""


# 爬取数据条目过滤
def item_filter(result, words):
    """
    根据关键字来过滤条目
    :param result: 爬取的结果集list(tuple)
    :param words: 关键字列表
    :return: 过滤的结果集
    """
    res = []
    if len(result) > 0 and len(words) > 0:
        for item in result:
            (title, href, dt) = item
            for key_word in words:
                if key_word in title:
                    res.append(item)
                    break
    return res
